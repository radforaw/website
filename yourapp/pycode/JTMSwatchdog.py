#!/usr/bin/env python
# -*- coding: utf-8 -*-
#
#  JTMSwatchdog.py
#  
#  Copyright 2019  <pi@raspberrypi>
#  
#  This program is free software; you can redistribute it and/or modify
#  it under the terms of the GNU General Public License as published by
#  the Free Software Foundation; either version 2 of the License, or
#  (at your option) any later version.
#  
#  This program is distributed in the hope that it will be useful,
#  but WITHOUT ANY WARRANTY; without even the implied warranty of
#  MERCHANTABILITY or FITNESS FOR A PARTICULAR PURPOSE.  See the
#  GNU General Public License for more details.
#  
#  You should have received a copy of the GNU General Public License
#  along with this program; if not, write to the Free Software
#  Foundation, Inc., 51 Franklin Street, Fifth Floor, Boston,
#  MA 02110-1301, USA.
#  
#  
import andyconfig
import requests
import datetime
import os
import xml.etree.cElementTree as ET
from collections import defaultdict

class JTMS():
    def __init__(self,camera):
        self.data=self.getaldata(camera)
        
    def getaldata(self,camera):
        url='http://bcc.opendata.onl/UTMC Flow.xml'
        par={'ApiKey':os.environ['ALKEY'],'SCN':camera,'Latest':str((datetime.datetime.now()+datetime.timedelta(days=1)).date()),
            'Earliest': str((datetime.datetime.now()-datetime.timedelta(days=28)).date()), 'TS':'true'}
        #print par
        n=requests.get(url,params=par)
        #print n.content[:1000]
        root = ET.fromstring(n.content)
        res=[]
        for a in root.findall('Data'):
            for b in a.findall('Data'):
                res.append(datetime.datetime.strptime(b.attrib['Date'],"%Y-%m-%d %H:%M:%S"))
        return res
                
    def days(self):
        j=defaultdict(int)
        for n in self.data:
            j[datetime.datetime.strftime(n,"%Y/%m/%d")]+=1
        return [[n,j[n]] for n in sorted(j.keys())]

    def mostcommoninterval(self):
        j=defaultdict(int)
        for n in range(len(self.data)-1):
            j[int(((self.data[n+1]-self.data[n]).total_seconds())/60)]+=1
        return [round(((float(j[5])/len(self.data))*100),2),max(j,key=j.get)]
    

def JTMSlist():
    url='http://bcc.opendata.onl/UTMC Flow.xml'
    par={'ApiKey':os.environ['ALKEY']}
    n=requests.get(url,params=par)
    print n.content[:1000]
    root = ET.fromstring(n.content)
    ret=[]
    for a in root.findall('Flow'):
        for b in a.findall('SCN'):
            c=b.text
            if c[:4]=='JTMS':
                ret.append([c,a.find('Description').text])
    return ret

from openpyxl import Workbook
from openpyxl.styles import PatternFill
wb=Workbook()
ws=wb.active
ws.title="JTMS Status"
r,c=1,1
for n in ["SCN","Location","Last updated","% within 5 minutes","Reporting Interval"]:
    d=ws.cell(row=r,column=c,value=n)
    d.fill=PatternFill("solid",fgColor="ADD8E6")
    c+=1

for n in JTMSlist():
    r+=1
    c=1
    a=JTMS(n[0])
    
    j=a.mostcommoninterval()
    if j[0]>80:
        ffc="00FF00"
    elif j[0]>60:
        ffc="FFFF00"
    else:
        ffc="FF0000"
    if j[1]==5:
        fac="00FF00"
    else:
        fac="FF0000"
    m=int(((datetime.datetime.now()-a.data[-1]).total_seconds())/86400)
    if m==0:
        i="Up to date"
        fc="00FF00"
    else:
        i=str(m)+" days ago"
        fc="FF0000"
    for n in [n[0],n[1],i,str(j[0])+'%','Every '+str(j[1])+" minutes"]:
        d=ws.cell(row=r,column=c,value=n)
        d.fill=PatternFill("solid",fgColor="D3D3D3")
        if c==3:
            d.fill=PatternFill("solid",fgColor=fc)
        if c==4:
            d.fill=PatternFill("solid",fgColor=ffc)
        if c==5:
            d.fill=PatternFill("solid",fgColor=fac)
        c+=1

def as_text(value):
    if value is None:
        return ""
    return str(value)

for column_cells in ws.columns:
    length = max(len(as_text(cell.value)) for cell in column_cells)
    ws.column_dimensions[column_cells[0].column].width = length

wb.save("website/website/yourapp/static/JTMS.xlsx")

'''
import csv
with open("results.csv","w") as csvfile:
    rit=csv.writer(csvfile)
    rit.writerow(["Location","SCN","Last updated","% within 5 minutes","Reporting interval"])
    for n in JTMSlist():
        a=JTMS(n[0])
        j=a.mostcommoninterval()
        m=int(((datetime.datetime.now()-a.data[-1]).total_seconds())/86400)
        if m==0:
            i="Up to date"
        else:
            i=str(m)+" days ago"
        rit.writerow([n[0],n[1],i,str(j[0])+'%','every '+str(j[1])+" minutes"])
    
'''
